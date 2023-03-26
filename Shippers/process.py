from openpyxl import load_workbook

xiaomi_del_list = [' EU ', ' RU ', ' RUСТБ ', 'Xiaomi ', ' JP ']
samsung_del_list = [' KZ ', ' AE ', ' AH ', ' EU ', ' RUTH ',
                    ' Simfree ', ' KR ', ' KZEU ', ' CNINAH ',
                    'INAHMY ', ' IN ', ' KZAEZA ', ' TH ', ' RU ', ' KZAEEU ', '  ']


def text_file_order_list(price_file):
    with open(price_file, 'r', encoding='utf8') as f:
        file = f.read()
    our_price = int()
    apple_ord = dict()
    apple_order_list = file.split('\n')
    for i in apple_order_list:
        k = i.split(' ')
        if int(k[-1]) in range(10000, 20000):
            our_price = int(k[-1]) + 1500
        if int(k[-1]) in range(20000, 30000):
            our_price = int(k[-1]) + 2500
        if int(k[-1]) in range(30000, 40000):
            our_price = int(k[-1]) + 3000
        if int(k[-1]) in range(40000, 50000):
            our_price = int(k[-1]) + 3500
        if int(k[-1]) in range(50000, 100000):
            our_price = int(k[-1]) + 5000
        if int(k[-1]) in range(100000, 300000):
            our_price = int(k[-1]) + 8000
        y = [(" ".join(k[:-1]), our_price)]
        apple_ord.update(y)
        y.clear()
    return apple_ord


def excel_order_list(price_file, del_list):
    order_price = int()
    price_list = list()
    xiaomi_ord = dict()
    wb = load_workbook(price_file)
    ws = wb["Лист1"]
    rows = ws.max_row
    cols = ws.max_column - 1
    for i in range(2, rows):
        string = str()
        for j in range(1, cols + 1):
            cell = ws.cell(row=i, column=j)
            string = string.replace(",", "").replace("/", "")
            for t in del_list:
                string = string.replace(t, ' ')
            string = string + str(cell.value) + ' '
        price_list.append(string.strip(' ').split(' '))
    price_list.sort(key=lambda arr: int(arr[-1]))
    for k in price_list:
        if int(k[-1]) in range(4000, 8000):
            order_price = int(k[-1]) + 1500
        if int(k[-1]) in range(8000, 15000):
            order_price = int(k[-1]) + 2500
        if int(k[-1]) in range(15000, 20000):
            order_price = int(k[-1]) + 3000
        if int(k[-1]) in range(20000, 30000):
            order_price = int(k[-1]) + 3500
        if int(k[-1]) in range(30000, 50000):
            order_price = int(k[-1]) + 5000
        if int(k[-1]) in range(50000, 300000):
            order_price = int(k[-1]) + 8000
        y = [(" ".join(k[:-1]), order_price)]
        xiaomi_ord.update(y)
    price_list.clear()
    for key, value in xiaomi_ord.items():
        price_list.append((str(key) + ' - ' + str(value)))
    return price_list
